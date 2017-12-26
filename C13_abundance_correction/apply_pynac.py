#!/usr/bin/env python2
# -*- coding:UTF-8 -*-
__author__ = 'Zhengtao Xiao'
###############################################
#usage: python apply_pynac.py config.txt
###############################################

from openpyxl import load_workbook
import sys,re,os
from datetime import datetime
import ConfigParser
import tempfile
import csv

#read config
parser = ConfigParser()
parser.read(sys.argv[1])
conf = {}
conf["pynacDir"] = parser.get("options","pynacDir")
conf["inFile"] = parser.get("options","inFile")
conf["formulaCol"] = parser.getint("options","formulaCol")
conf["formulaPattern"] = re.compile(parser.get("options","formulaPattern"))
conf["intensityCols"] = eval(parser.get("options","intensityCols"))
conf["outFile"] = parser.get("options","outFile")
#add pynac path
sys.path.append(conf["pynacDir"])
from pynac.analysis.analysis import Analysis


#read the file		
inWb = load_workbook(conf["inFile"])
sheet = inWb.active
max_ncol = sheet.max_column
max_nrow = sheet.max_row
sys.stderr.write("There are: %i rows and %i columns\n" % (max_nrow,max_ncol))

#read and extract the data
outDataDict = {}
header = ["GroupKey","intensity","13C-Count","Formula"]
for i in conf["intensityCols"]:
	outDataDict[i] = []
	outDataDict[i].append(header)

with tempfile.TemporaryDirectory() as tmpdir:
	sys.stderr.write("The temp files will be stored in %s\n" % tmpdir.name)
	group_id = 1
	for row in sheet.rows:
		if not str(row[0].value):
			sys.stderr.write("Error: please delete the blank rows")
			sys.exit(1)
		if row[0].row == 1:
			group_beg = 2
			continue
		if row[0].row == 2:
			formula = row[conf["formulaCol"]-1]
			c13Count = "0"
		if row[0].row >2:
			m = re.match(conf["formulaPattern"],row[conf["formularCol"]-1].strip())
			if m:
				fromula = m.group("formula")
				c13Count = m.group("C13_count")
			else:
				formula = row[conf["formulaCol"]-1]
				c13Count = "0"
				group_id += 1
		groupKey = "group_" + group_id
		for col_i in conf["intensityCols"]:
			outDataDict[col_i].append([groupKey,str(row[col_i-1]),c13Count,formular])
	#write into files and generate config for pynac
	finalData = {}
	for col_i in conf["intensityCols"]:
		outFile = os.path.join(tmpdir.name,str(col_i)+"_pynac_input.csv")
		with open(outFile,"w") as fout:
			for i in outDataDict[col_i]:
				fout.write(",".join(i)+"\n")
		#generate config file
		configFile = os.path.join(tmpdir.name,str(col_i)+"_pynac.conf")
		resultFile = os.path.join(tmpdir.name,str(col_i)+"_pynac_result.csv")
		with open(configFile,"w") as fout:
			fout.write('[PyNACOptions]' + "\n")
			fout.write("DataFile:" + outFile + "\n")
			fout.write('FormulaColumn: 3' + "\n")
			fout.write("Groupcolumns: [0]" + "\n")
			fout.write('IsotopeColumns: {"13C":2}' + "\n")
			fout.write('IntensityColumn: 1' + "\n")
			fout.write('SkipRows: [0]' + "\n")
			fout.write('OutputFile: ' + resultFile + "\n")
		#run pynac
		analysis = Analysis()
		analysis.ConfigureFromFile(configFile)
		analysis.Initialize()
		analysis.CorrectAll()
		analysis.WriteOutput()

		#get the corrected data from pynac file
		finalData[col_i] = []
		correct_column = 4
		with open(resultFile) as fin:
			for line in fin:
				tmp = line.strip().split(",")
				finalData.append(tmp[correct_column-1])
		
	#combine the pynac result into one file
	with open(conf["outFile"],"w") as fout:
		for i in range(len(finalData[col_i])):
			outList = [finalData[j][i] for j in conf["intensityCols"]]
			fout.write(",".join(outList) + "\n")

