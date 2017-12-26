#!/usr/bin/env python
# -*- coding:UTF-8 -*-
__author__ = 'Zhengtao Xiao'
###############################################
# the column named "Name" will be used for grouping molecules
# all the columns at right side of "MZ" columns will be calculated
# change 1: naming rule has been changed to: [13C]1TG(12:0/16:0/18:0)
################################################
from openpyxl import load_workbook
from openpyxl.styles import Border,Side
import sys
import re
import os

##please define these settings##
pattern = re.compile(r'^\[13C\]\d+') # for [13C]1TG(12:0/16:0/18:0)
#pattern = re.compile(r'^\d+\[13C\]') # for 1[13C]TG(12:0/16:0/18:0)
c13_abundance_percet = 0.1
#####################
def border_range(sheet,row_range,col_range,color="0000FF"):
	"""
	apply styles to a range of cells
	"""
	top = Border(top=Side(border_style="thin",color=color))
	right = Border(right=Side(border_style="thin",color=color))
	left = Border(left=Side(border_style="thin",color=color))
	bottom = Border(bottom=Side(border_style="thin",color=color))
	top_left = Border(top=Side(border_style="thin",color=color),
				left=Side(border_style="thin",color=color))
	top_right = Border(top=Side(border_style="thin",color=color),
				right=Side(border_style="thin",color=color))
	bot_left = Border(bottom=Side(border_style="thin",color=color),
				left=Side(border_style="thin",color=color))
	bot_right = Border(bottom=Side(border_style="thin",color=color),
				right=Side(border_style="thin",color=color))		
	#first line cell
	for c in col_range:
		sheet.cell(row=row_range[0],column=c).border = top
	#last line
	for c in col_range:
		sheet.cell(row=row_range[-1],column=c).border = bottom
	#left
	for r in row_range:
		sheet.cell(row=r,column=col_range[0]).border = left
		if r == row_range[0]:
			sheet.cell(row=r,column=col_range[0]).border = top_left
		if r == row_range[-1]:
			sheet.cell(row=r,column=col_range[0]).border = bot_left

	#right 
	for r in row_range:
		sheet.cell(row=r,column=col_range[-1]).border = right
		if r == row_range[0]:
			sheet.cell(row=r,column=col_range[-1]).border = top_right
		if r == row_range[-1]:
			sheet.cell(row=r,column=col_range[-1]).border = bot_right

#loading the file
inF = raw_input("please enter the input filename:\n")
if '.xls' not in inF:
	if os.path.exists(inF + ".xls"):
		inF += ".xls"
	elif os.path.exists(inF + ".xlsx"):
		inF += ".xlsx"
		
outname = raw_input("please enter the output filename\n")
if '.xls' not in outname:
	outname += '.xlsx'
if 'xls' not in inF:
	print "print check the file type"
	sys.exit()

inWb = load_workbook(inF,guess_types=True)
sheet = inWb.active
max_ncol = sheet.max_column
max_nrow = sheet.max_row
print "There are: %i rows and %i columns" % (max_nrow,max_ncol)
#find Elemental Composition and MZ columns
header_row = sheet.rows.next()
#header_row = sheet.rows[0]#on my notbook :
if not header_row[-1].value:
	print "Error columns number, please check !"
	sys.exit()
group_column = -1
data_begin = -1
for i in header_row:
	if str(i.value).strip().upper() == "NAME":
		group_column = i.col_idx
	if str(i.value).strip() == 'MZ':
		data_begin = i.col_idx +1
if group_column == -1:
	group_column = int(raw_input("Cannot find the \"Name\" column, please input:\n"))
if data_begin == -1:
	data_begin = int(raw_input("Please input the column number where data begins:\n"))

#fill the headline
for cell in header_row:
	if cell.col_idx >= data_begin:
		sheet.cell(row=1,column=cell.col_idx + max_ncol - data_begin +1, value=cell.value)
#add the nature abundance
#sheet.cell(row=1,column=cell.col_idx + max_ncol - data_begin +2,value="C13 nature abundance")

sum_begin_row = 2
for row in sheet.rows:
	if not str(row[group_column-1].value):
		print "Error: please delete the blank rows"
		sys.exit()
	name = str(row[group_column-1].value).strip()
	if (row[0].row > 2 and not re.match(pattern,name)) or row[0].row == max_nrow:
		if row[0].row == max_nrow :
			sum_end_row=row[0].row
		else:
			sum_end_row = row[0].row -1
		#fill the calculate
		row_range = range(sum_begin_row,sum_end_row+1)
		for column_i in xrange(max_ncol+1,2*max_ncol-data_begin+2):
			ref_column_i = column_i - max_ncol + data_begin -1
			for row_i in row_range:
				tmp_cell_coordinate = sheet.cell(row=row_i,column=ref_column_i).coordinate
				tmp_beg = sheet.cell(row=sum_begin_row,column=ref_column_i).coordinate
				tmp_end = sheet.cell(row=sum_end_row,column=ref_column_i).coordinate
				tmp_string = '=100*'+tmp_cell_coordinate+'/sum('+tmp_beg+":"+tmp_end+")"
				sheet.cell(row=row_i,column=column_i,value=tmp_string)
		#calculate the C13 natural abundance (wrong):
		# for r_idx,row_i in enumerate(row_range):
		# 	tmp_string = '=power(%f,%i)*power(1-%f,%i)' % (c13_abundance_percet,r_idx,c13_abundance_percet,len(row_range)-r_idx-1)
		# 	sheet.cell(row=row_i,column=2*max_ncol-data_begin+2,value=tmp_string)
		sum_begin_row = row[0].row
#add the border, make it more clearly.
border_range(sheet,range(1,max_nrow+1),range(max_ncol+1,2*max_ncol-data_begin+2))
#end ^_^
inWb.save(outname)

